# app.py
from __future__ import annotations

from datetime import date, timedelta
from typing import Any, Dict, List, Optional
import io

import pandas as pd
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from qf3_api import QF3Client, QF3Config


st.set_page_config(page_title="공정검사 이력조회", layout="wide")


def _to_date_str(d: date) -> str:
    return d.strftime("%Y-%m-%d")


def _df_from_rows(rows: List[Dict[str, Any]], cols: Optional[List[str]] = None) -> pd.DataFrame:
    df = pd.DataFrame(rows or [])
    if cols:
        for c in cols:
            if c not in df.columns:
                df[c] = None
        df = df[cols]
    return df


def _workcenter_from_head(h: dict) -> str:
    wc = (h.get("workcenterName") or "").strip()
    if wc:
        return wc
    ed = (h.get("equipmentDisplay") or "").strip()
    if ed:
        return ed.split("/")[0].strip()
    return (h.get("createdBy") or "").strip()


def _parse_and_tokens(q: str) -> List[str]:
    """
    품목명 검색:
      - "A%B%C" => ["A","B","C"] (AND)
      - 빈 토큰은 제거
    """
    q = (q or "").strip()
    if not q:
        return []
    parts = [p.strip() for p in q.split("%")]
    tokens = [p for p in parts if p]
    return tokens


def _filter_heads_by_item_name(heads: list, item_name_q: str) -> list:
    """
    heads(list[dict])에서 itemName에 대해 AND 포함검색 적용
    """
    tokens = _parse_and_tokens(item_name_q)
    if not tokens:
        return heads

    out = []
    for h in heads:
        name = (h.get("itemName") or "")
        name_l = name.lower()
        ok = True
        for t in tokens:
            if t.lower() not in name_l:
                ok = False
                break
        if ok:
            out.append(h)
    return out


def _excel_one_sheet_item_then_results(heads2: list, client: QF3Client, progress_cb=None) -> bytes:
    """
    한 시트 구조:
      - 품목 행 1줄(구분=품목)
      - 바로 아래에 해당 품목 검사결과(구분=검사) 여러 줄
      - 다음 품목 반복
    ※ 요청 반영: 품목 사이 '완전 빈 행'은 넣지 않음
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "품목_검사결과"

    columns = [
        "구분", "품목코드", "품목명", "작업장",
        "검사항목", "측정값", "단위", "기준", "상한", "하한", "판정"
    ]
    ws.append(columns)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center")
    for col_idx in range(1, len(columns) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    item_fill = PatternFill("solid", fgColor="2F2F2F")
    item_font = Font(bold=True, color="FFFFFF")
    item_align = Alignment(horizontal="left", vertical="center")

    row_num = 2
    total = len(heads2)

    for i, h in enumerate(heads2, start=1):
        wc = _workcenter_from_head(h)
        item_code = h.get("itemCode")
        item_name = h.get("itemName")
        mid = int(h["mfgInspectionId"])

        # 품목행
        ws.append(["품목", item_code, item_name, wc, "", "", "", "", "", "", ""])
        for col_idx in range(1, len(columns) + 1):
            c = ws.cell(row=row_num, column=col_idx)
            c.fill = item_fill
            c.font = item_font
            c.alignment = item_align
        row_num += 1

        # 검사결과 (바로 아래)
        resp_l = client.line_list(mfg_inspection_id=mid, page=1, limit=500, node="root")
        rows_l = client._extract_list(resp_l)

        if not rows_l:
            ws.append(["검사", item_code, item_name, wc, "(검사결과 없음)", "", "", "", "", "", ""])
            row_num += 1
        else:
            for r in rows_l:
                ws.append([
                    "검사",
                    item_code,
                    item_name,
                    wc,
                    r.get("level3ClassName"),
                    r.get("checkValue"),
                    r.get("unit"),
                    r.get("standardValue"),
                    r.get("upperLimit"),
                    r.get("lowerLimit"),
                    r.get("passDecision"),
                ])
                row_num += 1

        # ✅ 요청 반영: 품목 단위로 넣던 '완전 빈 줄' 제거
        # ws.append(["", "", "", "", "", "", "", "", "", "", ""])
        # row_num += 1

        if progress_cb:
            progress_cb(i, total)

    # 열 너비
    widths = [7, 14, 28, 14, 22, 10, 8, 10, 10, 10, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def init_state():
    ss = st.session_state
    ss.setdefault("logged_in", False)
    ss.setdefault("client", None)

    ss.setdefault("head_df", pd.DataFrame())
    ss.setdefault("head_rows", [])
    ss.setdefault("selected_mfgInspectionId", None)

    ss.setdefault("line_df", pd.DataFrame())


init_state()

st.title("공정검사 이력조회")

# ---------------- Login ----------------
with st.expander("로그인", expanded=not st.session_state.logged_in):
    c1, c2, c3, c4 = st.columns([1.2, 1.2, 1.5, 1.0])
    with c1:
        company_code = st.text_input("회사코드(companyCode)", value="")
    with c2:
        user_key = st.text_input("아이디(userKey)", value="")
    with c3:
        password = st.text_input("비밀번호(password)", value="", type="password")
    with c4:
        st.write("")
        st.write("")
        if st.button("로그인", use_container_width=True):
            try:
                if not company_code or not user_key or not password:
                    raise RuntimeError("회사코드/아이디/비밀번호를 입력하세요.")

                cfg = QF3Config(company_id=100, plant_id=11, language_code="KO")
                client = QF3Client(cfg)
                client.login(company_code=company_code, user_key=user_key, password=password)

                st.session_state.client = client
                st.session_state.logged_in = True
                st.success("로그인 성공")
            except Exception as e:
                st.session_state.logged_in = False
                st.session_state.client = None
                st.error(f"로그인 실패: {e}")

if not st.session_state.logged_in:
    st.stop()

client: QF3Client = st.session_state.client

# ---------------- Filters ----------------
st.subheader("조회")

# ✅ 품목명 검색칸 추가로 컬럼 하나 늘림
f1, f2, f3, f4, f5, f6, f7 = st.columns([1.1, 1.25, 1.25, 1.2, 1.05, 1.2, 1.75])
with f1:
    st.selectbox("공장", ["본사,주덕공장"], index=0, disabled=True)
with f2:
    d_from = st.date_input("검사일자 From", value=date.today())
with f3:
    d_to = st.date_input("검사일자 To", value=date.today())
with f4:
    job_name = st.text_input("작업지시(jobName)", value="")
with f5:
    operation_code = st.text_input("공정코드(operationCode)", value="")
with f6:
    item_code = st.text_input("품목코드(itemCode)", value="")
with f7:
    item_name_q = st.text_input("품목명(itemName)  (예: AR%NNB = AND검색)", value="")

g1, g2, g3 = st.columns([1.8, 1.2, 7.0])
with g1:
    fetch_all = st.checkbox("조회 결과 전체 로딩(페이지 제한 없이)", value=True)
with g2:
    limit = st.selectbox("페이지당", [20, 50, 100, 200, 500], index=0, disabled=fetch_all)
with g3:
    cbtn1, cbtn2 = st.columns([1.0, 1.2])
    with cbtn1:
        reset = st.button("초기화", use_container_width=True)
    with cbtn2:
        do_query = st.button("조회", type="primary", use_container_width=True)

if reset:
    st.session_state.head_df = pd.DataFrame()
    st.session_state.head_rows = []
    st.session_state.selected_mfgInspectionId = None
    st.session_state.line_df = pd.DataFrame()
    st.rerun()

inspection_date_from = _to_date_str(d_from)
inspection_date_to = _to_date_str(d_to)

released_from = _to_date_str(d_from - timedelta(days=7))
released_to = _to_date_str(d_to + timedelta(days=7))

if do_query:
    try:
        job_mp = client.build_job_equipment_map(released_date_from=released_from, released_date_to=released_to)

        # API에 item_name 파라미터가 확실치 않아서,
        # 1) 서버 조회는 기존처럼 가져오고
        # 2) 품목명은 로컬에서 AND 필터 적용
        if fetch_all:
            heads = client.fetch_all_heads(
                inspection_date_from=inspection_date_from,
                inspection_date_to=inspection_date_to,
                item_code=item_code,
                item_name="",
                job_name=job_name,
                operation_code=operation_code,
                person_id=0,
                check_class="OPR",
                limit=500,
            )
        else:
            resp = client.head_list(
                inspection_date_from=inspection_date_from,
                inspection_date_to=inspection_date_to,
                item_code=item_code,
                item_name="",
                job_name=job_name,
                operation_code=operation_code,
                person_id=0,
                check_class="OPR",
                page=1,
                limit=int(limit),
                start=1,
            )
            heads = client._extract_list(resp)

        heads2 = client.attach_equipment_to_heads(heads, job_mp)

        # ✅ 품목명 AND 필터 적용 (예: AR%NNB)
        heads2 = _filter_heads_by_item_name(heads2, item_name_q)

        head_cols = [
            "jobName", "operationNum", "operationCode",
            "itemCode", "itemName", "lotCode",
            "inspectionDate", "personName",
            "equipmentDisplay",
            "mfgInspectionName",
            "mfgInspectionId",
        ]
        head_df = _df_from_rows(heads2, head_cols)

        st.session_state.head_rows = heads2
        st.session_state.head_df = head_df
        st.session_state.selected_mfgInspectionId = None
        st.session_state.line_df = pd.DataFrame()

    except Exception as e:
        st.error(f"조회 실패: {e}")

# ---------------- Head Table ----------------
st.markdown("### 공정검사 이력")

head_df: pd.DataFrame = st.session_state.head_df
if head_df is None or head_df.empty:
    st.info("조회 조건 입력 후 [조회]를 누르세요.")
else:
    st.caption(f"표시 건수: {len(head_df):,}")

    evt = st.dataframe(
        head_df.drop(columns=["mfgInspectionId"], errors="ignore"),
        use_container_width=True,
        hide_index=True,
        on_select="rerun",
        selection_mode="single-row",
    )

    selected_rows = (evt.selection.rows if evt and hasattr(evt, "selection") else [])
    if selected_rows:
        idx = selected_rows[0]
        try:
            sel_id = int(st.session_state.head_df.iloc[idx]["mfgInspectionId"])
            st.session_state.selected_mfgInspectionId = sel_id
        except Exception:
            st.session_state.selected_mfgInspectionId = None

# ---------------- Line Table ----------------
st.markdown("### 검사 결과(선택 행)")

sel_id = st.session_state.selected_mfgInspectionId
if not sel_id:
    st.info("상단에서 행을 클릭하면 아래에 검사결과가 표시됩니다.")
else:
    try:
        resp2 = client.line_list(mfg_inspection_id=int(sel_id), page=1, limit=500, node="root")
        line_rows = client._extract_list(resp2)

        line_cols = [
            "level2ClassName", "level3ClassName",
            "checkMethod", "checkDevice",
            "checkValue", "unit",
            "standardValue", "upperLimit", "lowerLimit",
            "passDecision",
            "createdBy", "creationDate",
        ]
        line_df = _df_from_rows(line_rows, line_cols)
        st.session_state.line_df = line_df

        st.dataframe(line_df, use_container_width=True, hide_index=True)

    except Exception as e:
        st.error(f"검사결과 조회 실패: {e}")

# ---------------- Excel Download (ONE SHEET, FULL) ----------------
st.markdown("### 엑셀 다운로드 (한 시트 / 전체)")

if st.button("엑셀 다운로드(조회조건 전체)", use_container_width=True):
    try:
        # ✅ 엑셀은 항상 서버에서 '전체' 다시 긁어서 만든다
        job_mp = client.build_job_equipment_map(released_date_from=released_from, released_date_to=released_to)

        heads = client.fetch_all_heads(
            inspection_date_from=inspection_date_from,
            inspection_date_to=inspection_date_to,
            item_code=item_code,
            item_name="",
            job_name=job_name,
            operation_code=operation_code,
            person_id=0,
            check_class="OPR",
            limit=500,
        )
        heads2 = client.attach_equipment_to_heads(heads, job_mp)

        # ✅ 조회화면과 동일하게 품목명 AND 필터도 반영
        heads2 = _filter_heads_by_item_name(heads2, item_name_q)

        if not heads2:
            raise RuntimeError("다운로드할 데이터가 없습니다. 조회조건을 확인하세요.")

        prog = st.progress(0)

        def _cb(i, total):
            prog.progress(min(1.0, i / max(1, total)))

        xbytes = _excel_one_sheet_item_then_results(heads2, client, progress_cb=_cb)
        prog.empty()

        filename = f"공정검사_품목별검사결과_{inspection_date_from}_{inspection_date_to}.xlsx"
        st.download_button(
            label="⬇️ 다운로드",
            data=xbytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.success("엑셀 준비 완료")

    except Exception as e:
        st.error(f"엑셀 생성 실패: {e}")
